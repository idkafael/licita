import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy
import io
import re


def aplicar_reducao_planilha_licitacao(arquivo_bytes: bytes, percentual: float) -> bytes:
    """
    Aplica redução percentual na planilha de licitação.
    Reduz o 'Valor Unit sem BDI' de cada item e recalcula 'Valor Unit com BDI' e totais.
    """
    wb = load_workbook(io.BytesIO(arquivo_bytes))

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        col_sem_bdi = None
        col_bdi = None
        col_com_bdi = None
        col_total_sem = None
        col_total_com = None
        col_quant = None
        header_row = None

        # Detectar cabeçalhos
        for row in ws.iter_rows(min_row=1, max_row=30):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip().upper()
                    if any(x in val for x in ['VALOR UNIT', 'PREÇO', 'PRECO', 'P.UNIT', 'P. UNIT']):
                        if 'SEM BDI' in val or 'S/ BDI' in val or ('SEM' in val and 'BDI' in val):
                            col_sem_bdi = cell.column
                            header_row = cell.row
                        elif 'COM BDI' in val or 'C/ BDI' in val or ('COM' in val and 'BDI' in val):
                            col_com_bdi = cell.column
                        elif val in ['BDI', 'B.D.I', 'B.D.I.']:
                            col_bdi = cell.column
                    elif val in ['BDI', 'B.D.I', 'B.D.I.'] and col_bdi is None:
                        col_bdi = cell.column
                    elif 'TOTAL' in val and 'SEM' in val:
                        col_total_sem = cell.column
                    elif 'TOTAL' in val and 'COM' in val:
                        col_total_com = cell.column
                    elif 'TOTAL' in val and col_total_com is None and col_total_sem is None:
                        col_total_com = cell.column
                    elif val in ['QUANT', 'QUANT.', 'QUANTIDADE', 'QTD', 'QTD.']:
                        col_quant = cell.column

        if col_sem_bdi is None:
            # Tentar detecção pela coluna PREÇO genérica (R$)
            for row in ws.iter_rows(min_row=1, max_row=30):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        val = cell.value.strip().upper()
                        if val in ['SEM BDI', 'UNITÁRIO', 'UNITARIO', 'UNIT.']:
                            col_sem_bdi = cell.column
                            header_row = cell.row

        if col_sem_bdi is None:
            continue  # Pular aba se não encontrar

        fator = 1 - (percentual / 100)

        # Aplicar nas linhas de dados
        for row in ws.iter_rows(min_row=(header_row or 1) + 1):
            for cell in row:
                if cell.column == col_sem_bdi:
                    if isinstance(cell.value, (int, float)) and cell.value and cell.value > 0:
                        novo_valor_sem_bdi = cell.value * fator
                        cell.value = round(novo_valor_sem_bdi, 2)

                        # Recalcular valor com BDI
                        if col_bdi and col_com_bdi:
                            bdi_cell = ws.cell(row=cell.row, column=col_bdi)
                            com_bdi_cell = ws.cell(row=cell.row, column=col_com_bdi)
                            if isinstance(bdi_cell.value, (int, float)) and bdi_cell.value:
                                com_bdi_cell.value = round(novo_valor_sem_bdi * (1 + bdi_cell.value), 2)
                            elif not isinstance(com_bdi_cell.value, str):
                                # Se não tem BDI separado, manter fórmula ou calcular sem
                                pass

                        # Recalcular totais
                        if col_quant:
                            quant_cell = ws.cell(row=cell.row, column=col_quant)
                            if isinstance(quant_cell.value, (int, float)) and quant_cell.value:
                                if col_total_sem:
                                    ws.cell(row=cell.row, column=col_total_sem).value = round(
                                        novo_valor_sem_bdi * quant_cell.value, 2
                                    )
                                if col_total_com and col_bdi:
                                    bdi_cell = ws.cell(row=cell.row, column=col_bdi)
                                    if isinstance(bdi_cell.value, (int, float)):
                                        ws.cell(row=cell.row, column=col_total_com).value = round(
                                            novo_valor_sem_bdi * (1 + bdi_cell.value) * quant_cell.value, 2
                                        )
                                elif col_total_com and col_com_bdi:
                                    com_bdi_cell = ws.cell(row=cell.row, column=col_com_bdi)
                                    if isinstance(com_bdi_cell.value, (int, float)):
                                        ws.cell(row=cell.row, column=col_total_com).value = round(
                                            com_bdi_cell.value * quant_cell.value, 2
                                        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def aplicar_reducao_cpu(arquivo_bytes: bytes, percentual: float) -> bytes:
    """
    Aplica redução percentual na aba CPU (Composição de Preço Unitário).
    Reduz os preços unitários dos insumos.
    """
    wb = load_workbook(io.BytesIO(arquivo_bytes))

    for sheet_name in wb.sheetnames:
        if 'CPU' not in sheet_name.upper() and 'COMPOSI' not in sheet_name.upper():
            continue

        ws = wb[sheet_name]
        fator = 1 - (percentual / 100)

        col_punit = None
        col_ptotal = None
        col_quant = None

        # Detectar cabeçalhos na aba CPU
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip().upper()
                    if val in ['P.UNIT.', 'P. UNIT.', 'P.UNIT', 'PREÇO UNIT', 'PRECO UNIT', 'VALOR UNIT']:
                        col_punit = cell.column
                    elif val in ['P.TOTAL', 'P. TOTAL', 'P.TOTAL.', 'TOTAL', 'VALOR TOTAL']:
                        col_ptotal = cell.column
                    elif val in ['QUANT.', 'QUANT', 'QUANTIDADE', 'QTD']:
                        col_quant = cell.column

        if col_punit is None:
            continue

        for row in ws.iter_rows(min_row=1):
            for cell in row:
                if cell.column == col_punit:
                    if isinstance(cell.value, (int, float)) and cell.value and cell.value > 0:
                        novo_punit = cell.value * fator
                        cell.value = round(novo_punit, 2)

                        if col_ptotal and col_quant:
                            quant_cell = ws.cell(row=cell.row, column=col_quant)
                            if isinstance(quant_cell.value, (int, float)) and quant_cell.value:
                                ws.cell(row=cell.row, column=col_ptotal).value = round(
                                    novo_punit * quant_cell.value, 2
                                )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def aplicar_reducao_cronograma(arquivo_bytes: bytes, percentual: float) -> bytes:
    """
    Aplica redução no cronograma físico-financeiro.
    Reduz os valores monetários mantendo as proporções por mês.
    """
    wb = load_workbook(io.BytesIO(arquivo_bytes))

    for sheet_name in wb.sheetnames:
        if 'CRONOGRAMA' not in sheet_name.upper():
            continue

        ws = wb[sheet_name]
        fator = 1 - (percentual / 100)

        col_valor = None
        header_row = None

        for row in ws.iter_rows(min_row=1, max_row=20):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip().upper()
                    if val in ['VALOR (R$)', 'VALOR', 'TOTAL', 'R$', 'VALOR R$']:
                        col_valor = cell.column
                        header_row = cell.row

        if col_valor is None:
            continue

        for row in ws.iter_rows(min_row=(header_row or 1) + 1):
            for cell in row:
                if cell.column == col_valor:
                    if isinstance(cell.value, (int, float)) and cell.value and cell.value > 0:
                        cell.value = round(cell.value * fator, 2)
                # Reduzir também valores mensais (colunas após valor total)
                elif col_valor and cell.column > col_valor:
                    if isinstance(cell.value, (int, float)) and cell.value and cell.value > 0:
                        # Verificar se parece um valor monetário (não percentual)
                        if cell.value > 1:
                            cell.value = round(cell.value * fator, 2)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def aplicar_reducao_orcamento_resumido(arquivo_bytes: bytes, percentual: float) -> bytes:
    """
    Aplica redução no orçamento resumido.
    Reduz os valores totais por categoria.
    """
    wb = load_workbook(io.BytesIO(arquivo_bytes))

    fator = 1 - (percentual / 100)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        col_total = None
        header_row = None

        for row in ws.iter_rows(min_row=1, max_row=20):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip().upper()
                    if val in ['TOTAL', 'VALOR TOTAL', 'R$', 'VALOR (R$)', 'TOTAL (R$)']:
                        col_total = cell.column
                        header_row = cell.row

        if col_total is None:
            continue

        for row in ws.iter_rows(min_row=(header_row or 1) + 1):
            for cell in row:
                if cell.column == col_total:
                    if isinstance(cell.value, (int, float)) and cell.value and cell.value > 0:
                        cell.value = round(cell.value * fator, 2)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def processar_planilha(arquivo_bytes: bytes, tipo: str, percentual: float) -> bytes:
    """
    Dispatcher principal — escolhe a função certa pelo tipo de planilha.
    """
    tipo = tipo.lower()
    if tipo == 'licitacao':
        return aplicar_reducao_planilha_licitacao(arquivo_bytes, percentual)
    elif tipo == 'cpu':
        return aplicar_reducao_cpu(arquivo_bytes, percentual)
    elif tipo == 'cronograma':
        return aplicar_reducao_cronograma(arquivo_bytes, percentual)
    elif tipo == 'orcamento_resumido':
        return aplicar_reducao_orcamento_resumido(arquivo_bytes, percentual)
    else:
        return aplicar_reducao_planilha_licitacao(arquivo_bytes, percentual)


def preview_planilha(arquivo_bytes: bytes, tipo: str, percentual: float) -> dict:
    """
    Gera preview das alterações sem salvar.
    Retorna lista de itens com valor original e novo valor.
    """
    wb = load_workbook(io.BytesIO(arquivo_bytes))
    itens = []
    valor_original_total = 0
    valor_novo_total = 0

    fator = 1 - (percentual / 100)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        col_sem_bdi = None
        col_desc = None
        col_total = None
        header_row = None

        for row in ws.iter_rows(min_row=1, max_row=30):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip().upper()
                    if any(x in val for x in ['DISCRIMIN', 'DESCRI', 'SERVI']):
                        col_desc = cell.column
                    if 'SEM BDI' in val or 'S/ BDI' in val or val in ['P.UNIT.', 'P.UNIT', 'VALOR UNIT']:
                        col_sem_bdi = cell.column
                        header_row = cell.row
                    if 'TOTAL' in val and col_total is None:
                        col_total = cell.column

        if col_sem_bdi is None:
            continue

        count = 0
        for row in ws.iter_rows(min_row=(header_row or 1) + 1):
            valor_cell = None
            desc_cell = None

            for cell in row:
                if cell.column == col_sem_bdi:
                    valor_cell = cell
                if col_desc and cell.column == col_desc:
                    desc_cell = cell

            if valor_cell and isinstance(valor_cell.value, (int, float)) and valor_cell.value and valor_cell.value > 0:
                count += 1
                if count <= 50:  # Limitar preview a 50 itens
                    descricao = str(desc_cell.value) if desc_cell and desc_cell.value else f'Item {count}'
                    if len(descricao) > 60:
                        descricao = descricao[:57] + '...'
                    itens.append({
                        'descricao': descricao,
                        'valor_original': round(valor_cell.value, 2),
                        'valor_novo': round(valor_cell.value * fator, 2),
                        'reducao': round(valor_cell.value - (valor_cell.value * fator), 2)
                    })
                valor_original_total += valor_cell.value
                valor_novo_total += valor_cell.value * fator

    return {
        'itens': itens,
        'total_itens': len(itens),
        'valor_original_total': round(valor_original_total, 2),
        'valor_novo_total': round(valor_novo_total, 2),
        'reducao_total': round(valor_original_total - valor_novo_total, 2),
        'percentual': percentual
    }
